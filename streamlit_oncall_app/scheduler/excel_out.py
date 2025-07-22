import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.worksheet.page import PageMargins

def to_excel(df, path, title="当直表"):
    """df -> Excel (A4横、罫線入り)。path は str でも BytesIO でもOK"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Schedule"

    ws.merge_cells("A1:E1")
    ws["A1"] = title
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")


    headers = ["日付", "曜", "枠1", "枠2", "備考"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")


    for r, row in enumerate(df.itertuples(index=False), start=3):
        ws.cell(row=r, column=1, value=row.date.strftime("%m/%d"))
        ws.cell(row=r, column=2, value=row.weekday)
        ws.cell(row=r, column=3, value=row.slot1)
        ws.cell(row=r, column=4, value=row.slot2)
        ws.cell(row=r, column=5, value="")  # 備考


    thin = Side(style="thin", color="000000")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=5):
        for cell in row:
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
            cell.alignment = Alignment(horizontal="center", vertical="center")


    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 4
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 20

    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.print_title_rows = "1:2"
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)

    if isinstance(path, str):
        wb.save(path)
    else:
        wb.save(path)
